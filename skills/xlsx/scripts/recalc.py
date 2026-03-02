"""
Excel 公式重算脚本
使用 LibreOffice 重新计算 Excel 文件中的所有公式
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from office.soffice import get_soffice_env

from openpyxl import load_workbook

MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def has_gtimeout():
    """检查系统是否安装了 gtimeout 命令"""
    try:
        subprocess.run(
            ["gtimeout", "--version"], capture_output=True, timeout=1, check=False
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def setup_libreoffice_macro():
    """设置 LibreOffice 宏，用于重算并保存 Excel 文件"""
    macro_dir = os.path.expanduser(
        MACRO_DIR_MACOS if platform.system() == "Darwin" else MACRO_DIR_LINUX
    )
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    # 如果宏文件已存在且包含 RecalculateAndSave 函数，则无需重新创建
    if (
        os.path.exists(macro_file)
        and "RecalculateAndSave" in Path(macro_file).read_text()
    ):
        return True

    # 如果宏目录不存在，先启动 LibreOffice 初始化配置，然后创建目录
    if not os.path.exists(macro_dir):
        subprocess.run(
            ["soffice", "--headless", "--terminate_after_init"],
            capture_output=True,
            timeout=10,
            env=get_soffice_env(),
        )
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """
    使用 LibreOffice 重新计算 Excel 文件中的所有公式

    参数:
        filename: Excel 文件路径
        timeout: 超时时间（秒），默认 30 秒

    返回:
        包含状态和错误详情的字典
    """
    if not Path(filename).exists():
        return {"error": f"文件 {filename} 不存在"}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {"error": "无法设置 LibreOffice 宏"}

    cmd = [
        "soffice",
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    # Linux 使用 timeout 命令，macOS 使用 gtimeout（如果可用）
    if platform.system() == "Linux":
        cmd = ["timeout", str(timeout)] + cmd
    elif platform.system() == "Darwin" and has_gtimeout():
        cmd = ["gtimeout", str(timeout)] + cmd

    result = subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env())

    # 返回码 124 表示超时，其他非零返回码表示错误
    if result.returncode != 0 and result.returncode != 124:
        error_msg = result.stderr or "重算过程中发生未知错误"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice 宏未正确配置"}
        return {"error": error_msg}

    try:
        wb = load_workbook(filename, data_only=True)

        # Excel 常见错误类型列表
        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        # 遍历所有工作表和单元格，检查错误
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        # 汇总各类型错误的详情（最多显示 20 个位置）
        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        # 统计公式总数
        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count

        return result

    except Exception as e:
        return {"error": str(e)}


def main():
    """命令行入口函数"""
    if len(sys.argv) < 2:
        print("用法: python recalc.py <excel文件> [超时秒数]")
        print("\n使用 LibreOffice 重新计算 Excel 文件中的所有公式")
        print("\n返回包含错误详情的 JSON:")
        print("  - status: 'success' 或 'errors_found'")
        print("  - total_errors: 发现的 Excel 错误总数")
        print("  - total_formulas: 文件中的公式数量")
        print("  - error_summary: 按错误类型分类的详情及位置")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
